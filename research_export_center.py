"""Research export utilities for the 3alimnIA evaluator workspace.

The module keeps export preparation independent from Streamlit widgets so it can
be validated with small, deterministic tests.  It provides strict anonymisation
support, analysis-ready filters, styled Excel workbooks, data dictionaries,
and reproducibility bundles with SHA-256 manifests.
"""

from __future__ import annotations

import hashlib
import io
import json
import re
import zipfile
from dataclasses import asdict, dataclass
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, Mapping, Optional, Sequence, Tuple

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


DIRECT_IDENTIFIER_COLUMNS = {
    "full_name",
    "email",
    "institution",
    "password_hash",
    "raw_name",
    "name",
    "student_id",
    "evaluator_username",
}

# These fields may contain names or other identifiers typed by a participant or
# evaluator.  They are omitted from the default research-safe export.  The full
# administrative backup remains the protected location for these fields.
FREE_TEXT_RISK_COLUMNS = {
    "prompt",
    "response",
    "diagnostic",
    "reflection_text",
    "attempt_text",
    "open_feedback_json",
    "overall_comment",
    "event_detail",
    "consent_text",
    "selected_text",
}

COMMON_COLUMN_DESCRIPTIONS = {
    "participant_code": "Stable pseudonymous participant code used to link datasets.",
    "study_group": "Study-arm assignment or single-arm marker.",
    "academic_level": "Reported academic level (for example Licence, Master, or PhD).",
    "preferred_language": "Interface/content language selected by the learner.",
    "pre_score": "Pre-test score on the platform scale.",
    "post_score": "Post-test score on the platform scale.",
    "learning_gain": "Post-test minus pre-test score, in percentage points.",
    "progress_percent": "Percentage of required learning-study steps completed.",
    "is_complete_case": "Whether the participant completed the full study protocol.",
    "ai_interactions": "Number of logged AI tutor interactions.",
    "created_at": "UTC timestamp recorded when the row/event was created.",
    "updated_at": "UTC timestamp recorded when the row was last updated.",
    "interaction_id": "Identifier of an AI tutor interaction within the source system.",
    "mode": "AI support mode, such as external LLM or rule-based fallback.",
    "provider": "AI service provider used for the interaction.",
    "model": "Model identifier reported by the provider.",
    "latency_ms": "End-to-end response latency in milliseconds.",
    "response_word_count": "Number of words in the generated response.",
    "pedagogical_quality_score": "Aggregate LPQS pedagogical quality score.",
    "conceptual_accuracy": "LPQS conceptual-accuracy criterion score.",
    "answer_relevance": "LPQS answer-relevance criterion score.",
    "pedagogical_clarity": "LPQS pedagogical-clarity criterion score.",
    "scaffolding_quality": "LPQS scaffolding-quality criterion score.",
    "qiskit_alignment": "LPQS Qiskit-alignment criterion score.",
    "reflection_support": "LPQS reflection-support criterion score.",
    "personalization": "LPQS personalization criterion score.",
    "seconds_before_ai": "Elapsed seconds between the learning action and AI request.",
    "event_type": "Machine-readable event type recorded by the platform.",
    "lesson_id": "Stable lesson identifier.",
    "question_id": "Stable assessment-question identifier.",
    "is_correct": "Whether the recorded answer was correct.",
}


@dataclass(frozen=True)
class ExportFilters:
    study_groups: Tuple[str, ...] = ()
    academic_levels: Tuple[str, ...] = ()
    completion: str = "all"  # all | complete | incomplete
    date_from: Optional[str] = None
    date_to: Optional[str] = None

    def to_dict(self) -> Dict[str, Any]:
        return asdict(self)


def utc_stamp() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")


def sha256_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def safe_filename(value: str, default: str = "export") -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", str(value or "").strip()).strip("._")
    return cleaned or default


def _normalise_date(value: Any) -> Optional[pd.Timestamp]:
    if value in (None, ""):
        return None
    parsed = pd.to_datetime(value, errors="coerce", utc=True)
    if pd.isna(parsed):
        return None
    return parsed


def _participant_codes_from_progress(progress: pd.DataFrame, filters: ExportFilters) -> Optional[set[str]]:
    if progress is None or progress.empty or "participant_code" not in progress.columns:
        return None
    filtered = progress.copy()
    if filters.study_groups and "study_group" in filtered.columns:
        filtered = filtered[filtered["study_group"].astype(str).isin(filters.study_groups)]
    if filters.academic_levels and "academic_level" in filtered.columns:
        filtered = filtered[filtered["academic_level"].astype(str).isin(filters.academic_levels)]
    if filters.completion == "complete" and "is_complete_case" in filtered.columns:
        filtered = filtered[filtered["is_complete_case"].astype(bool)]
    elif filters.completion == "incomplete" and "is_complete_case" in filtered.columns:
        filtered = filtered[~filtered["is_complete_case"].astype(bool)]
    return set(filtered["participant_code"].dropna().astype(str).tolist())


def filter_export_tables(
    tables: Mapping[str, pd.DataFrame],
    filters: ExportFilters,
) -> Dict[str, pd.DataFrame]:
    """Apply participant-level and date filters consistently across datasets."""
    progress = tables.get("progress_summary", pd.DataFrame())
    allowed_codes = _participant_codes_from_progress(progress, filters)
    start = _normalise_date(filters.date_from)
    end = _normalise_date(filters.date_to)
    if end is not None:
        end = end.normalize() + pd.Timedelta(days=1) - pd.Timedelta(microseconds=1)

    filtered_tables: Dict[str, pd.DataFrame] = {}
    for name, frame in tables.items():
        if frame is None:
            filtered_tables[name] = pd.DataFrame()
            continue
        out = frame.copy()
        if allowed_codes is not None and "participant_code" in out.columns:
            out = out[out["participant_code"].astype(str).isin(allowed_codes)]

        date_col = next((c for c in ("created_at", "updated_at", "date", "timestamp") if c in out.columns), None)
        if date_col and (start is not None or end is not None) and not out.empty:
            parsed = pd.to_datetime(out[date_col], errors="coerce", utc=True)
            mask = parsed.notna()
            if start is not None:
                mask &= parsed >= start
            if end is not None:
                mask &= parsed <= end
            out = out[mask]
        filtered_tables[name] = out.reset_index(drop=True)
    return filtered_tables


def _safe_cell(value: Any) -> Any:
    if isinstance(value, (dict, list, tuple, set)):
        value = json.dumps(value, ensure_ascii=False, default=str)
    if isinstance(value, pd.Timestamp):
        value = value.isoformat()
    if isinstance(value, datetime):
        value = value.isoformat()
    if isinstance(value, date):
        value = value.isoformat()
    if isinstance(value, str) and len(value) > 32000:
        return value[:31970] + "…[truncated]"
    return value


def excel_safe_frame(frame: pd.DataFrame) -> pd.DataFrame:
    if frame is None:
        return pd.DataFrame()
    out = frame.copy()
    for col in out.columns:
        if pd.api.types.is_datetime64_any_dtype(out[col]):
            out[col] = pd.to_datetime(out[col], errors="coerce").map(lambda x: x.isoformat() if not pd.isna(x) else None)
        elif out[col].dtype == "object":
            out[col] = out[col].map(_safe_cell)
    return out


def _clean_sheet_name(name: str, used: set[str]) -> str:
    base = re.sub(r"[\\/*?:\[\]]", "_", str(name))[:31] or "Sheet"
    candidate = base
    suffix = 1
    while candidate in used:
        tail = f"_{suffix}"
        candidate = f"{base[:31-len(tail)]}{tail}"
        suffix += 1
    used.add(candidate)
    return candidate


def tables_to_excel_bytes(
    tables: Mapping[str, pd.DataFrame],
    *,
    metadata: Optional[Mapping[str, Any]] = None,
) -> bytes:
    """Create a styled, analysis-ready Excel workbook."""
    output = io.BytesIO()
    used: set[str] = set()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        if metadata:
            meta = pd.DataFrame([{"field": key, "value": _safe_cell(value)} for key, value in metadata.items()])
            meta.to_excel(writer, sheet_name="README", index=False)
            used.add("README")
        for dataset, frame in tables.items():
            sheet = _clean_sheet_name(dataset, used)
            excel_safe_frame(frame).to_excel(writer, sheet_name=sheet, index=False)

        header_fill = PatternFill("solid", fgColor="143B66")
        header_font = Font(color="FFFFFF", bold=True)
        for ws in writer.book.worksheets:
            ws.freeze_panes = "A2"
            if ws.max_row >= 1 and ws.max_column >= 1:
                ws.auto_filter.ref = ws.dimensions
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            for column_cells in ws.columns:
                letter = get_column_letter(column_cells[0].column)
                values = [str(cell.value or "") for cell in column_cells[:250]]
                max_len = max([len(x) for x in values] + [8])
                ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 48)
            ws.sheet_view.showGridLines = False
    return output.getvalue()


def _sensitivity_for_column(column: str) -> str:
    lower = str(column).lower()
    if lower in DIRECT_IDENTIFIER_COLUMNS:
        return "direct_identifier"
    if lower in FREE_TEXT_RISK_COLUMNS:
        return "free_text_risk"
    if "code" in lower and lower != "participant_code":
        return "internal_identifier"
    return "research_variable"


def dataset_inventory(tables: Mapping[str, pd.DataFrame]) -> pd.DataFrame:
    rows = []
    for name, frame in tables.items():
        df = frame if frame is not None else pd.DataFrame()
        date_col = next((c for c in ("created_at", "updated_at", "date") if c in df.columns), None)
        min_date = max_date = None
        if date_col and not df.empty:
            parsed = pd.to_datetime(df[date_col], errors="coerce", utc=True).dropna()
            if not parsed.empty:
                min_date = parsed.min().isoformat()
                max_date = parsed.max().isoformat()
        rows.append(
            {
                "dataset": name,
                "rows": int(len(df)),
                "columns": int(len(df.columns)),
                "participants": int(df["participant_code"].nunique()) if "participant_code" in df.columns else None,
                "date_min": min_date,
                "date_max": max_date,
                "contains_direct_identifier": bool(set(df.columns) & DIRECT_IDENTIFIER_COLUMNS),
                "contains_free_text_risk": bool(set(df.columns) & FREE_TEXT_RISK_COLUMNS),
            }
        )
    return pd.DataFrame(rows)


def data_dictionary(tables: Mapping[str, pd.DataFrame]) -> pd.DataFrame:
    rows = []
    for dataset, frame in tables.items():
        df = frame if frame is not None else pd.DataFrame()
        for column in df.columns:
            series = df[column]
            rows.append(
                {
                    "dataset": dataset,
                    "column": column,
                    "dtype": str(series.dtype),
                    "non_null": int(series.notna().sum()),
                    "missing": int(series.isna().sum()),
                    "unique": int(series.nunique(dropna=True)) if len(series) else 0,
                    "sensitivity": _sensitivity_for_column(column),
                    "description": COMMON_COLUMN_DESCRIPTIONS.get(
                        str(column), str(column).replace("_", " ").strip().capitalize()
                    ),
                }
            )
    return pd.DataFrame(rows)


def add_derived_analysis_tables(tables: Mapping[str, pd.DataFrame]) -> Dict[str, pd.DataFrame]:
    out = {name: frame.copy() for name, frame in tables.items()}
    progress = out.get("progress_summary", pd.DataFrame())
    if not progress.empty:
        paired = progress.dropna(subset=[c for c in ("pre_score", "post_score") if c in progress.columns]).copy()
        if {"pre_score", "post_score"}.issubset(paired.columns):
            paired["learning_gain"] = pd.to_numeric(paired["post_score"], errors="coerce") - pd.to_numeric(paired["pre_score"], errors="coerce")
        out["paired_scores"] = paired
        summary = {
            "n_participants": int(len(progress)),
            "n_complete_cases": int(progress["is_complete_case"].astype(bool).sum()) if "is_complete_case" in progress else None,
            "n_paired_pre_post": int(len(paired)),
            "mean_pre": float(pd.to_numeric(paired.get("pre_score"), errors="coerce").mean()) if not paired.empty and "pre_score" in paired else None,
            "mean_post": float(pd.to_numeric(paired.get("post_score"), errors="coerce").mean()) if not paired.empty and "post_score" in paired else None,
            "mean_gain": float(pd.to_numeric(paired.get("learning_gain"), errors="coerce").mean()) if not paired.empty and "learning_gain" in paired else None,
        }
        out["paper_summary"] = pd.DataFrame([summary])
        if "study_group" in progress.columns:
            numeric = [c for c in ("pre_score", "post_score", "learning_gain", "progress_percent", "ai_interactions") if c in progress.columns]
            agg: Dict[str, Any] = {"participant_code": "count"} if "participant_code" in progress.columns else {}
            agg.update({c: "mean" for c in numeric})
            if agg:
                group = progress.groupby("study_group", dropna=False).agg(agg).reset_index()
                group = group.rename(columns={"participant_code": "participants"})
                out["group_summary"] = group

    surveys = out.get("surveys", pd.DataFrame())
    if not surveys.empty and "responses_json" in surveys.columns:
        rows = []
        for _, row in surveys.iterrows():
            try:
                values = json.loads(row.get("responses_json") or "{}")
            except Exception:
                values = {}
            base = {"participant_code": row.get("participant_code"), "created_at": row.get("created_at")}
            rows.append({**base, **{f"survey_{k}": v for k, v in values.items()}})
        out["survey_item_scores"] = pd.DataFrame(rows)
    return out


def infer_date_bounds(tables: Mapping[str, pd.DataFrame]) -> Tuple[Optional[date], Optional[date]]:
    timestamps = []
    for frame in tables.values():
        if frame is None or frame.empty:
            continue
        for col in ("created_at", "updated_at", "date"):
            if col in frame.columns:
                parsed = pd.to_datetime(frame[col], errors="coerce", utc=True).dropna()
                if not parsed.empty:
                    timestamps.extend([parsed.min(), parsed.max()])
                break
    if not timestamps:
        return None, None
    return min(timestamps).date(), max(timestamps).date()


def codebook_markdown(tables: Mapping[str, pd.DataFrame], filters: ExportFilters, anonymized: bool) -> str:
    inventory = dataset_inventory(tables)
    lines = [
        "# 3alimnIA Research Export Codebook",
        "",
        f"- Generated (UTC): {datetime.now(timezone.utc).isoformat()}",
        f"- Export mode: {'strict anonymized research export' if anonymized else 'protected administrative backup'}",
        f"- Filters: `{json.dumps(filters.to_dict(), ensure_ascii=False)}`",
        "",
        "## Privacy interpretation",
        "",
        "The research workbook uses a stable participant code and excludes direct account identifiers.",
        "Free-text fields that can contain self-disclosed identifiers are excluded from the strict anonymized export.",
        "This is a research-safe pseudonymous dataset; the participant-code mapping must remain access-controlled.",
        "",
        "## Datasets",
        "",
    ]
    for _, row in inventory.iterrows():
        lines.append(f"### {row['dataset']}")
        lines.append(f"Rows: {row['rows']} · Columns: {row['columns']} · Participants: {row['participants']}")
        lines.append("")
    return "\n".join(lines)


def analysis_template_py(workbook_name: str) -> str:
    return f'''"""Starter analysis for a 3alimnIA research export."""
from pathlib import Path
import pandas as pd

WORKBOOK = Path(__file__).with_name("{workbook_name}")
all_tables = pd.read_excel(WORKBOOK, sheet_name=None)
print("Available datasets:", list(all_tables))

progress = all_tables.get("progress_summary", pd.DataFrame())
paired = all_tables.get("paired_scores", pd.DataFrame())

if not paired.empty and {{"pre_score", "post_score"}}.issubset(paired.columns):
    paired = paired.dropna(subset=["pre_score", "post_score"]).copy()
    paired["learning_gain"] = paired["post_score"] - paired["pre_score"]
    print(paired[["pre_score", "post_score", "learning_gain"]].describe())

    try:
        from scipy import stats
        result = stats.ttest_rel(paired["post_score"], paired["pre_score"], nan_policy="omit")
        print("Paired t-test:", result)
    except ImportError:
        print("Install scipy to run inferential tests: pip install scipy")
'''


def build_reproducibility_bundle(
    tables: Mapping[str, pd.DataFrame],
    *,
    filters: ExportFilters,
    app_version: str,
    anonymized: bool = True,
    prefix: str = "3alimnia_research",
) -> Tuple[bytes, Dict[str, Any]]:
    """Return a ZIP bundle and its manifest."""
    stamp = utc_stamp()
    workbook_name = f"{safe_filename(prefix)}_{stamp}.xlsx"
    metadata = {
        "app_version": app_version,
        "generated_utc": datetime.now(timezone.utc).isoformat(),
        "export_mode": "strict_anonymized" if anonymized else "full_administrative",
        "filters": json.dumps(filters.to_dict(), ensure_ascii=False),
    }
    workbook = tables_to_excel_bytes(tables, metadata=metadata)
    inventory = dataset_inventory(tables)
    dictionary = data_dictionary(tables)
    codebook = codebook_markdown(tables, filters, anonymized)
    readme = (
        "3alimnIA reproducibility bundle\n\n"
        f"Workbook: {workbook_name}\n"
        "CSV copies are stored in the csv/ directory.\n"
        "Use data_dictionary.csv and codebook.md before statistical analysis.\n"
        "The strict research export excludes direct identifiers and risky raw free text.\n"
        "Keep the participant-code mapping outside this bundle and under separate access control.\n"
    )

    files: Dict[str, bytes] = {
        workbook_name: workbook,
        "dataset_inventory.csv": inventory.to_csv(index=False).encode("utf-8-sig"),
        "data_dictionary.csv": dictionary.to_csv(index=False).encode("utf-8-sig"),
        "codebook.md": codebook.encode("utf-8"),
        "analysis_template.py": analysis_template_py(workbook_name).encode("utf-8"),
        "README.txt": readme.encode("utf-8"),
    }
    for name, frame in tables.items():
        files[f"csv/{safe_filename(name)}.csv"] = excel_safe_frame(frame).to_csv(index=False).encode("utf-8-sig")

    hashes = {name: sha256_bytes(data) for name, data in files.items()}
    manifest = {
        "app_version": app_version,
        "generated_utc": datetime.now(timezone.utc).isoformat(),
        "anonymized": anonymized,
        "filters": filters.to_dict(),
        "datasets": dataset_inventory(tables).to_dict(orient="records"),
        "files": hashes,
    }
    files["manifest.json"] = json.dumps(manifest, ensure_ascii=False, indent=2, default=str).encode("utf-8")
    files["SHA256SUMS.txt"] = "\n".join(f"{digest}  {name}" for name, digest in sorted(hashes.items())).encode("utf-8")

    output = io.BytesIO()
    with zipfile.ZipFile(output, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for name, data in files.items():
            archive.writestr(name, data)
    return output.getvalue(), manifest


def export_ui_copy(lang: str) -> Dict[str, str]:
    copies = {
        "ar": {
            "filters": "نطاق التصدير",
            "group": "مجموعة الدراسة",
            "level": "المستوى الأكاديمي",
            "completion": "حالة الاكتمال",
            "all": "الكل",
            "complete": "الحالات المكتملة فقط",
            "incomplete": "الحالات غير المكتملة فقط",
            "date_filter": "تقييد التصدير بفترة زمنية",
            "from": "من",
            "to": "إلى",
            "prepare": "تحضير الحزمة البحثية",
            "strict_privacy": "حماية الخصوصية الصارمة مفعلة: تُحذف المعرّفات المباشرة والنصوص الحرة عالية المخاطر.",
            "inventory": "جرد مجموعات البيانات",
            "dictionary": "قاموس البيانات",
            "preview": "معاينة البيانات",
            "download_workbook": "تنزيل Excel البحثي",
            "download_bundle": "تنزيل حزمة إعادة الإنتاج ZIP",
            "download_csv": "تنزيل الجدول المحدد CSV",
            "dataset": "مجموعة البيانات",
            "rows": "الصفوف",
            "participants": "المشاركون",
            "hash": "بصمة SHA-256",
            "admin_title": "نسخة إدارية كاملة",
            "admin_warning": "تتضمن بيانات تعريفية ونصوصًا خامًا. احفظها في مساحة مؤمنة ولا تشاركها مع فريق التحليل.",
            "admin_confirm": "أؤكد أنني أحتاج نسخة إدارية محمية",
            "prepare_admin": "تحضير النسخة الإدارية",
            "audit": "سجل عمليات التصدير",
            "no_data": "لا توجد بيانات ضمن النطاق المحدد.",
        },
        "fr": {
            "filters": "Périmètre d’export",
            "group": "Groupe d’étude",
            "level": "Niveau académique",
            "completion": "Statut d’achèvement",
            "all": "Tous",
            "complete": "Cas complets uniquement",
            "incomplete": "Cas incomplets uniquement",
            "date_filter": "Limiter à une période",
            "from": "Du",
            "to": "Au",
            "prepare": "Préparer le dossier de recherche",
            "strict_privacy": "Protection stricte active : identifiants directs et textes libres à risque sont exclus.",
            "inventory": "Inventaire des jeux de données",
            "dictionary": "Dictionnaire des données",
            "preview": "Aperçu",
            "download_workbook": "Télécharger le classeur Excel",
            "download_bundle": "Télécharger le paquet reproductible ZIP",
            "download_csv": "Télécharger le tableau CSV",
            "dataset": "Jeu de données",
            "rows": "Lignes",
            "participants": "Participants",
            "hash": "Empreinte SHA-256",
            "admin_title": "Sauvegarde administrative complète",
            "admin_warning": "Contient des identifiants et du texte brut. Stockez-la dans un espace protégé.",
            "admin_confirm": "Je confirme avoir besoin d’une sauvegarde administrative protégée",
            "prepare_admin": "Préparer la sauvegarde",
            "audit": "Journal des exports",
            "no_data": "Aucune donnée dans le périmètre sélectionné.",
        },
        "en": {
            "filters": "Export scope",
            "group": "Study group",
            "level": "Academic level",
            "completion": "Completion status",
            "all": "All",
            "complete": "Complete cases only",
            "incomplete": "Incomplete cases only",
            "date_filter": "Restrict to a date range",
            "from": "From",
            "to": "To",
            "prepare": "Prepare research package",
            "strict_privacy": "Strict privacy is active: direct identifiers and high-risk free text are excluded.",
            "inventory": "Dataset inventory",
            "dictionary": "Data dictionary",
            "preview": "Data preview",
            "download_workbook": "Download research Excel",
            "download_bundle": "Download reproducibility ZIP",
            "download_csv": "Download selected CSV",
            "dataset": "Dataset",
            "rows": "Rows",
            "participants": "Participants",
            "hash": "SHA-256 fingerprint",
            "admin_title": "Full administrative backup",
            "admin_warning": "Contains identifying data and raw text. Store it in a protected location and do not share it with the analysis team.",
            "admin_confirm": "I confirm that I need a protected administrative backup",
            "prepare_admin": "Prepare administrative backup",
            "audit": "Export audit log",
            "no_data": "No data are available for the selected scope.",
        },
    }
    return copies.get(lang, copies["en"])
